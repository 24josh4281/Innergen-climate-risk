"""
make_excel_site.py — 임의 위도/경도의 전체 기후 리스크 데이터를 Excel로 저장

사용법:
    python make_excel_site.py 36.0095 129.3435 "포항공장"
    python make_excel_site.py 37.5649 126.9793 "OCI_HQ"

출력: site_output/{이름}_{lat}_{lon}_climate_risk.xlsx
"""
import sys
import os
import asyncio
import logging
from pathlib import Path
from datetime import datetime

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
API_DIR  = BASE_DIR / "api"
OUT_DIR  = BASE_DIR / "site_output"
OUT_DIR.mkdir(exist_ok=True)

sys.path.insert(0, str(API_DIR))
os.chdir(API_DIR)

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── 인자 파싱 ──────────────────────────────────────────────────────────────────
lat  = float(sys.argv[1]) if len(sys.argv) > 1 else 36.0095
lon  = float(sys.argv[2]) if len(sys.argv) > 2 else 129.3435
name = sys.argv[3]        if len(sys.argv) > 3 else f"{lat}_{lon}"

# ── 모듈 초기화 ────────────────────────────────────────────────────────────────
logger.info("데이터 초기화 중...")
from data_loader import site_data
from cmip6_grid import cmip6_grid
site_data.load()
cmip6_grid.load()

from tier_resolver import resolve, SSP_KEYS, PERIOD_KEYS, PERIOD_LABELS

# ── 데이터 조회 ────────────────────────────────────────────────────────────────
logger.info(f"좌표 조회: {lat}°N, {lon}°E")
result = asyncio.run(resolve(lat, lon))

drivers = result["drivers"]
meta    = result["meta"]

res = meta.get('resolution', '?')
tier_label = {"precomputed": "T1", "1deg": "T2", "2deg": "T3"}.get(res, res)
if meta.get("kma_rda"):
    tier_label += "+KMA_RDA"
if meta.get("kma_cordex"):
    tier_label += "+KMA_CORDEX"
logger.info(f"Tier: {tier_label} | 해상도: {res}")

# ── openpyxl 임포트 ────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("openpyxl 미설치. 설치: pip install openpyxl")
    sys.exit(1)

# ── 스타일 헬퍼 ────────────────────────────────────────────────────────────────
DARK_BG   = "0B1628"
CARD_BG   = "111B2E"
HEADER_BG = "1B2E4A"
ACCENT    = "4FC3F7"
RED_BG    = "C62828";  RED_FG    = "FFFFFF"
AMBER_BG  = "F57F17";  AMBER_FG  = "000000"
GREEN_BG  = "1B5E20";  GREEN_FG  = "FFFFFF"
TEXT_FG   = "E2EAF4"
DIM_FG    = "7B9CC4"
GROUP_BG  = "162236"

def hfont(bold=False, color=TEXT_FG, size=10):
    return Font(name="Consolas", bold=bold, color=color, size=size)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hborder():
    s = Side(style="thin", color="1E3250")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value, bold=False, fg=TEXT_FG, bg=CARD_BG, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = hfont(bold=bold, color=fg)
    c.fill      = hfill(bg)
    c.border    = hborder()
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if num_fmt:
        c.number_format = num_fmt
    return c

def rag_style(ws, row, col, value, rag):
    if rag == "red":
        bg, fg = RED_BG, RED_FG
    elif rag == "amber":
        bg, fg = AMBER_BG, AMBER_FG
    else:
        bg, fg = GREEN_BG, GREEN_FG
    return cell_style(ws, row, col, value, fg=fg, bg=bg, align="center")

def header_row(ws, row, values, bg=HEADER_BG):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = hfont(bold=True, color=ACCENT)
        c.fill      = hfill(bg)
        c.border    = hborder()
        c.alignment = Alignment(horizontal="center", vertical="center")

def group_header(ws, row, label, ncols, bg=GROUP_BG):
    c = ws.cell(row=row, column=1, value=label)
    c.font      = hfont(bold=True, color=ACCENT, size=9)
    c.fill      = hfill(bg)
    c.border    = hborder()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 16

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell

def autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        maxlen = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(maxlen + 2, min_w), max_w)

# ── 값 추출 헬퍼 ──────────────────────────────────────────────────────────────
def get_val(entry):
    if entry is None:
        return None
    if isinstance(entry, dict):
        v = entry.get("value")
    else:
        v = entry
    if v is None:
        return None
    return round(float(v), 3) if isinstance(v, (int, float)) else v

def fmt(v, decimals=1):
    if v is None:
        return "N/A"
    if isinstance(v, float):
        return round(v, decimals)
    return v

# ── 드라이버 그룹 정의 ────────────────────────────────────────────────────────
GROUPS = [
    ("CMIP6 기후변수", [
        ("tasmax",  "최고기온",          "°C"),
        ("tasmin",  "최저기온",          "°C"),
        ("tas",     "평균기온",          "°C"),
        ("pr",      "강수량",            "mm/day"),
        ("prsn",    "적설량",            "mm/day"),
        ("sfcWind", "지표풍속",          "m/s"),
        ("evspsbl", "증발산량",          "mm/day"),
    ]),
    ("ETCCDI 극값 — 열", [
        ("etccdi_txx",  "월최고기온(TXx)",      "°C"),
        ("etccdi_tnn",  "월최저기온(TNn)",      "°C"),
        ("etccdi_su",   "여름일수(SU>25°C)",    "days/yr"),
        ("etccdi_tr",   "열대야(TR>20°C)",      "days/yr"),
        ("etccdi_fd",   "결빙일수(FD<0°C)",     "days/yr"),
        ("etccdi_wsdi", "온난기간(WSDI)",        "days/yr"),
        ("etccdi_wbgt", "습구흑구온도(WBGT)",    "°C"),
    ]),
    ("ETCCDI 극값 — 강수", [
        ("etccdi_cdd",    "연속건조일수(CDD)",   "days"),
        ("etccdi_cwd",    "연속습윤일수(CWD)",   "days"),
        ("etccdi_rx1day", "1일최대강수(Rx1day)", "mm"),
        ("etccdi_rx5day", "5일최대강수(Rx5day)", "mm"),
        ("etccdi_r95p",   "극한강수(R95p)",      "mm/yr"),
        ("etccdi_sdii",   "강수강도(SDII)",      "mm/wet-day"),
    ]),
    ("PhyRisk — 열·노동", [
        ("heat_stress",       "열 스트레스",       "score"),
        ("extreme_heat_35c",  "극한고온 35°C",     "score"),
        ("work_loss_high",    "노동손실(고강도)",  "score"),
        ("work_loss_medium",  "노동손실(중강도)",  "score"),
        ("heat_degree_days",  "냉방도일",          "score"),
    ]),
    ("PhyRisk — 수자원·가뭄", [
        ("water_stress",   "물 스트레스",    "score"),
        ("water_depletion","수자원 고갈",    "score"),
        ("drought_risk",   "가뭄 위험",      "score"),
    ]),
    ("PhyRisk — 홍수·태풍·기타", [
        ("flood_risk",     "홍수 위험",      "score"),
        ("river_flood",    "하천 홍수",      "score"),
        ("coastal_flood",  "해안 홍수",      "score"),
        ("pluvial_flood",  "도시 침수",      "score"),
        ("cyclone_risk",   "사이클론 위험",  "score"),
        ("storm_surge",    "폭풍 해일",      "score"),
        ("sea_level_rise", "해수면 상승",    "score"),
        ("wildfire_risk",  "산불 위험",      "score"),
        ("earthquake_risk","지진 위험",      "score"),
        ("landslide_risk", "산사태 위험",    "score"),
    ]),
    ("CLIMADA HDF5 — 정량 EAL (연간기대손실, 역사적 정적값)", [
        ("TC_EAL",       "태풍 EAL",    "score 0-100"),
        ("Flood_EAL",    "홍수 EAL",    "score 0-100"),
        ("Wildfire_EAL", "산불 EAL",    "score 0-100"),
        ("EQ_EAL",       "지진 EAL",    "score 0-100"),
    ]),
    ("Aqueduct 수자원 (정적)", [
        ("aq_water_stress",      "수자원 스트레스",      "0-5"),
        ("aq_river_flood",       "하천홍수 위험",        "0-5"),
        ("aq_coastal_flood",     "해안홍수 위험",        "0-5"),
        ("aq_drought",           "가뭄 위험",            "0-5"),
        ("aq_interann_var",      "연간변동성",           "0-5"),
        ("aq_water_stress_2050", "수자원 스트레스 2050", "0-5"),
    ]),
    ("IBTrACS 태풍 (역사적)", [
        ("tc_annual_freq",  "태풍 연간 빈도",    "건/yr"),
        ("tc_max_wind_kt",  "최대 풍속",         "kt"),
        ("tc_cat3_count",   "카테고리3+ 태풍",   "건"),
    ]),
    ("PSHA 지진재해 (정적)", [
        ("psha_pga_475",  "PGA 475년 재현주기",  "g"),
        ("psha_pga_2475", "PGA 2475년 재현주기", "g"),
    ]),
    ("CCKP 에너지·극한열 — World Bank 0.25°", [
        ("cckp_hi35",  "열지수 초과일 (HI>35°C)",  "days/yr"),
        ("cckp_hd40",  "극한고온일 (Tmax>40°C)",   "days/yr"),
        ("cckp_tr26",  "열대야 한국기준 (>26°C)",  "days/yr"),
        ("cckp_cdd65", "냉방도일 (CDD base 65°F)", "°F·day/yr"),
        ("cckp_hdd65", "난방도일 (HDD base 65°F)", "°F·day/yr"),
    ]),
    ("CCKP ETCCDI 교차검증 — 온도 계열 (World Bank 0.25°)", [
        ("cckp_csdi",       "한파 지속기간 (CSDI)",    "days/yr"),
        ("cckp_wsdi",       "온난 지속기간 (WSDI-CP)", "days/yr"),
        ("cckp_cdd_consec", "연속건조일수 (CDD-CP)",   "days"),
    ]),
    ("CCKP 신규 확장 — 가뭄·습도·강수·한랭 (World Bank 0.25°)", [
        ("cckp_spei12",  "가뭄지수 (SPEI-12)",       "index"),
        ("cckp_gsl",     "생장기간 (GSL)",            "days/yr"),
        ("cckp_hurs",    "상대습도 (HURS)",           "%"),
        ("cckp_id",      "결빙일 (Tmax<0°C)",        "days/yr"),
        ("cckp_rxmonth", "월최대강수 (Rx-month)",     "mm/month"),
    ]),
    ("KMA 농촌진흥청 ENS — 167개 행정구역 통계적 상세화 (SSP 4종)", [
        ("kma_tasmax",  "최고기온 (KMA_RDA)",   "°C"),
        ("kma_tasmin",  "최저기온 (KMA_RDA)",   "°C"),
        ("kma_tas",     "평균기온 (KMA_RDA)",   "°C"),
        ("kma_pr",      "강수량 (KMA_RDA)",     "mm/day"),
        ("kma_hurs",    "상대습도 (KMA_RDA)",   "%"),
        ("kma_sfcWind", "풍속 (KMA_RDA)",       "m/s"),
    ]),
    ("KMA CORDEX — MOHC-HadGEM2-ES/GERICS-REMO2015 동역학 상세화 (~22km)", [
        ("cordex_tasmax",  "최고기온 (CORDEX)",  "°C"),
        ("cordex_tasmin",  "최저기온 (CORDEX)",  "°C"),
        ("cordex_tas",     "평균기온 (CORDEX)",  "°C"),
        ("cordex_pr",      "강수량 (CORDEX)",    "mm/day"),
        ("cordex_sfcWind", "풍속 (CORDEX)",      "m/s"),
        ("cordex_hurs",    "상대습도 (CORDEX)",  "%"),
        ("cordex_rsds",    "일사량 (CORDEX)",    "MJ/m²/day"),
    ]),
]

SSP_LABELS = {
    "ssp126": "SSP1-2.6",
    "ssp245": "SSP2-4.5",
    "ssp370": "SSP3-7.0",
    "ssp585": "SSP5-8.5",
}
PERIOD_LABEL_SHORT = {
    "baseline": "현재",
    "near":     "단기(2030s)",
    "mid":      "중기(2050s)",
    "far":      "장기(2080s)",
    "end":      "말기(2090s)",
}
SSP_ORDER = ["ssp126", "ssp245", "ssp370", "ssp585"]

# ── RAG 임계값 (간단 버전) ────────────────────────────────────────────────────
RAG_RULES = {
    "tasmax":         lambda v: "red" if v > 38 else "amber" if v > 33 else "green",
    "tasmin":         lambda v: "amber" if v > 22 else "green",
    "tas":            lambda v: "red" if v > 30 else "amber" if v > 25 else "green",
    "pr":             lambda v: "amber" if v > 10 or v < 1 else "green",
    "prsn":           lambda v: "green",
    "sfcWind":        lambda v: "red" if v > 15 else "amber" if v > 10 else "green",
    "evspsbl":        lambda v: "amber" if v > 5 else "green",
    "TC_EAL":         lambda v: "red" if v > 50 else "amber" if v > 20 else "green",
    "Flood_EAL":      lambda v: "red" if v > 50 else "amber" if v > 20 else "green",
    "Wildfire_EAL":   lambda v: "red" if v > 50 else "amber" if v > 20 else "green",
    "EQ_EAL":         lambda v: "red" if v > 50 else "amber" if v > 20 else "green",
    "etccdi_txx":     lambda v: "red" if v > 38 else "amber" if v > 33 else "green",
    "etccdi_tnn":     lambda v: "amber" if v < -10 else "green",
    "etccdi_su":      lambda v: "red" if v > 100 else "amber" if v > 60 else "green",
    "etccdi_tr":      lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "etccdi_fd":      lambda v: "amber" if v > 60 else "green",
    "etccdi_wsdi":    lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "etccdi_wbgt":    lambda v: "red" if v > 32 else "amber" if v > 28 else "green",
    "etccdi_cdd":     lambda v: "red" if v > 60 else "amber" if v > 30 else "green",
    "etccdi_cwd":     lambda v: "green",
    "etccdi_rx1day":  lambda v: "red" if v > 100 else "amber" if v > 50 else "green",
    "etccdi_rx5day":  lambda v: "red" if v > 200 else "amber" if v > 100 else "green",
    "etccdi_r95p":    lambda v: "red" if v > 500 else "amber" if v > 200 else "green",
    "etccdi_sdii":    lambda v: "amber" if v > 20 else "green",
    "cckp_hi35":      lambda v: "red" if v > 60 else "amber" if v > 20 else "green",
    "cckp_hd40":      lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "cckp_tr26":      lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "cckp_cdd65":     lambda v: "red" if v > 2500 else "amber" if v > 1500 else "green",
    "cckp_hdd65":     lambda v: "red" if v > 4000 else "amber" if v > 2000 else "green",
    "cckp_csdi":      lambda v: "red" if v > 10 else "amber" if v > 4 else "green",
    "cckp_wsdi":      lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "cckp_cdd_consec":lambda v: "red" if v > 60 else "amber" if v > 30 else "green",
    "cckp_spei12":    lambda v: "red" if v < -1.5 else "amber" if v < -0.5 else "green",
    "cckp_gsl":       lambda v: "red" if v > 350 else "amber" if v > 320 else "green",
    "cckp_hurs":      lambda v: "red" if v > 80 else "amber" if v > 72 else "green",
    "cckp_id":        lambda v: "red" if v > 30 else "amber" if v > 10 else "green",
    "cckp_rxmonth":   lambda v: "red" if v > 400 else "amber" if v > 250 else "green",
}

def get_rag(dk, val):
    if val is None:
        return None
    fn = RAG_RULES.get(dk)
    if fn:
        return fn(val)
    # score 계열 (0~1)
    if isinstance(val, float) and 0 <= val <= 1:
        return "red" if val > 0.6 else "amber" if val > 0.3 else "green"
    # 0~5 계열
    if isinstance(val, (int, float)) and 0 <= val <= 5:
        return "red" if val >= 3 else "amber" if val >= 2 else "green"
    return "green"


# ══════════════════════════════════════════════════════════════════════════════
# Excel 생성
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # 기본 시트 제거

PERIOD_COLS = list(PERIOD_KEYS)  # ["baseline","near","mid","far","end"]


# ── 시트 1~4: SSP별 통합 ───────────────────────────────────────────────────────
for ssp in SSP_ORDER:
    ssp_data  = drivers.get(ssp, {})
    ws = wb.create_sheet(SSP_LABELS[ssp])
    ws.sheet_view.showGridLines = False

    # 배경
    ws.sheet_properties.tabColor = "1B2E4A"

    # 제목행
    r = 1
    c = ws.cell(row=r, column=1,
                value=f"기후 물리적 리스크 — {name}  |  {lat}°N, {lon}°E  |  {SSP_LABELS[ssp]}")
    c.font      = hfont(bold=True, color=ACCENT, size=11)
    c.fill      = hfill(DARK_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 22

    # 부제목 (메타)
    r = 2
    c = ws.cell(row=r, column=1,
                value=f"데이터 소스: CMIP6 17모델 앙상블 | PhyRisk | Aqueduct | IBTrACS | PSHA | CCKP    생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font      = hfont(bold=False, color=DIM_FG, size=9)
    c.fill      = hfill(DARK_BG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 14

    # 컬럼 헤더
    r = 3
    COL_HEADERS = ["카테고리", "변수명", "변수키", "단위"] + [PERIOD_LABEL_SHORT[p] for p in PERIOD_COLS]
    header_row(ws, r, COL_HEADERS)
    ws.row_dimensions[r].height = 18

    r = 4
    for grp_label, vars_list in GROUPS:
        group_header(ws, r, f"▶  {grp_label}", len(COL_HEADERS))
        r += 1
        for dk, var_label, unit in vars_list:
            vals = [get_val((ssp_data.get(p) or {}).get(dk)) for p in PERIOD_COLS]
            # 카테고리
            cell_style(ws, r, 1, grp_label.split(" ")[0], bg=CARD_BG, fg=DIM_FG)
            cell_style(ws, r, 2, var_label, bold=True, bg=CARD_BG)
            cell_style(ws, r, 3, dk, bg=CARD_BG, fg=DIM_FG)
            cell_style(ws, r, 4, unit, bg=CARD_BG, fg=DIM_FG, align="center")
            for ci, (p, v) in enumerate(zip(PERIOD_COLS, vals), 5):
                rag = get_rag(dk, v)
                if rag:
                    rag_style(ws, r, ci, fmt(v), rag)
                else:
                    cell_style(ws, r, ci, fmt(v), align="center")
            ws.row_dimensions[r].height = 15
            r += 1

    # 열 너비 고정
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    for ci in range(5, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    freeze(ws, "B4")


# ── 시트 5: 전체 통합비교 (변수 × SSP × 시점 한눈에) ───────────────────────────
ws = wb.create_sheet("전체_통합비교")
ws.sheet_view.showGridLines = False

# 열 구성: 카테고리(A) | 변수명(B) | 단위(C) | SSP126×5기간 | SSP245×5기간 | SSP370×5기간 | SSP585×5기간
LABEL_COLS  = 3          # A,B,C
PERIOD_CNT  = len(PERIOD_COLS)   # 5
TOTAL_COLS  = LABEL_COLS + len(SSP_ORDER) * PERIOD_CNT  # 3 + 20 = 23

# 제목행
r = 1
c = ws.cell(row=r, column=1,
            value=f"기후 물리적 리스크 전체 통합 — {name}  |  {lat}°N, {lon}°E  |  SSP 4종 × 5시점")
c.font      = hfont(bold=True, color=ACCENT, size=11)
c.fill      = hfill(DARK_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
ws.row_dimensions[r].height = 22

# SSP 그룹 헤더 (병합)
r = 2
for ci_label in range(1, LABEL_COLS + 1):
    c = ws.cell(row=r, column=ci_label, value="")
    c.fill   = hfill(HEADER_BG)
    c.border = hborder()
for si, ssp in enumerate(SSP_ORDER):
    col_start = LABEL_COLS + 1 + si * PERIOD_CNT
    col_end   = col_start + PERIOD_CNT - 1
    c = ws.cell(row=r, column=col_start, value=SSP_LABELS[ssp])
    c.font      = hfont(bold=True, color="FFFFFF", size=10)
    ssp_colors  = {"ssp126": "1B5E20", "ssp245": "1565C0", "ssp370": "E65100", "ssp585": "B71C1C"}
    c.fill      = PatternFill("solid", fgColor=ssp_colors[ssp])
    c.border    = hborder()
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
    for ci in range(col_start + 1, col_end + 1):
        ws.cell(row=r, column=ci).fill   = PatternFill("solid", fgColor=ssp_colors[ssp])
        ws.cell(row=r, column=ci).border = hborder()
ws.row_dimensions[r].height = 18

# 기간 서브헤더
r = 3
cell_style(ws, r, 1, "카테고리", bold=True, bg=HEADER_BG, fg=ACCENT, align="center")
cell_style(ws, r, 2, "변수명",   bold=True, bg=HEADER_BG, fg=ACCENT, align="center")
cell_style(ws, r, 3, "단위",     bold=True, bg=HEADER_BG, fg=ACCENT, align="center")
period_short = [PERIOD_LABEL_SHORT[p] for p in PERIOD_COLS]
for si in range(len(SSP_ORDER)):
    for pi, plabel in enumerate(period_short):
        col = LABEL_COLS + 1 + si * PERIOD_CNT + pi
        c = ws.cell(row=r, column=col, value=plabel)
        c.font      = hfont(bold=True, color=ACCENT, size=8)
        c.fill      = hfill(HEADER_BG)
        c.border    = hborder()
        c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 16

# 데이터 행
r = 4
for grp_label, vars_list in GROUPS:
    group_header(ws, r, f"▶  {grp_label}", TOTAL_COLS)
    r += 1
    for dk, var_label, unit in vars_list:
        row_bg = CARD_BG if r % 2 == 0 else "0E1C30"
        cell_style(ws, r, 1, grp_label.split(" ")[0], bg=row_bg, fg=DIM_FG)
        cell_style(ws, r, 2, var_label, bold=True, bg=row_bg)
        cell_style(ws, r, 3, unit, bg=row_bg, fg=DIM_FG, align="center")
        for si, ssp in enumerate(SSP_ORDER):
            ssp_data = drivers.get(ssp, {})
            for pi, p in enumerate(PERIOD_COLS):
                col = LABEL_COLS + 1 + si * PERIOD_CNT + pi
                v   = get_val((ssp_data.get(p) or {}).get(dk))
                rag = get_rag(dk, v)
                if rag:
                    rag_style(ws, r, col, fmt(v), rag)
                else:
                    cell_style(ws, r, col, fmt(v), bg=row_bg, align="center")
        ws.row_dimensions[r].height = 15
        r += 1

# 열 너비
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 11
for ci in range(LABEL_COLS + 1, TOTAL_COLS + 1):
    ws.column_dimensions[get_column_letter(ci)].width = 9

freeze(ws, "C4")


# ── 시트 6: 원시데이터 (Long format) ─────────────────────────────────────────
ws = wb.create_sheet("전체_원시데이터")
ws.sheet_view.showGridLines = False
ws.row_dimensions[1].height = 18
header_row(ws, 1, ["카테고리", "SSP", "시점", "변수키", "변수명", "단위", "값", "RAG"])

r = 2
for ssp in SSP_ORDER:
    ssp_data = drivers.get(ssp, {})
    for p in PERIOD_COLS:
        period_data = ssp_data.get(p, {})
        for grp_label, vars_list in GROUPS:
            for dk, var_label, unit in vars_list:
                v = get_val(period_data.get(dk))
                rag = get_rag(dk, v)
                cat = grp_label.split(" ")[0]
                row_vals = [cat, SSP_LABELS[ssp], PERIOD_LABEL_SHORT[p], dk, var_label, unit, fmt(v), rag or "N/A"]
                for ci, rv in enumerate(row_vals, 1):
                    bg = CARD_BG if r % 2 == 0 else "0E1C30"
                    cell_style(ws, r, ci, rv, bg=bg, align="left" if ci < 7 else "center")
                r += 1

autofit(ws)
freeze(ws, "E2")


# ── 시트 6: 분석 정보 ─────────────────────────────────────────────────────────
ws = wb.create_sheet("분석_정보")
ws.sheet_view.showGridLines = False

info_rows = [
    ("항목", "값"),
    ("분석 플랫폼", "Innergen Climate Scenario"),
    ("생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("위도 (Lat)", lat),
    ("경도 (Lon)", lon),
    ("사업장명", name),
    ("데이터 Tier", tier_label),
    ("", ""),
    ("데이터 소스", "설명"),
    ("CMIP6", "17개 모델 앙상블 — SSP126/245/370/585 × 5시점"),
    ("PhyRisk", "OS-Climate PhyRisk — 18개 위험 유형 (0~1 스케일)"),
    ("CLIMADA HDF5", "ETH Zürich CLIMADA — TC/홍수/산불/지진 EAL 4종 (역사적 정적값, 0~100 지수)"),
    ("Aqueduct", "WRI Aqueduct 4.0 — 수자원 위험 6종 (0~5 스케일)"),
    ("IBTrACS", "NOAA IBTrACS v04 — 역사적 태풍 통계 (1980-2023)"),
    ("PSHA", "GEM Global Earthquake Model — PGA (475yr, 2475yr)"),
    ("CCKP", "World Bank CCKP CMIP6 0.25° — 에너지·극한열 5종 + ETCCDI 3종 + 신규확장 5종"),
    ("KMA_RDA", "농촌진흥청 ENS 앙상블 — 한반도 167개 행정구역, SSP 4종 × 6변수 (통계적 상세화)"),
    ("KMA_CORDEX", "기상청 CORDEX 1km — 한반도 1km 격자, SSP 4종 × 27극한지수 (동역학적 상세화, 데이터 준비 시 활성화)"),
    ("", ""),
    ("기준", "값"),
    ("시나리오", "SSP1-2.6 / SSP2-4.5 / SSP3-7.0 / SSP5-8.5"),
    ("시점 (현재)", "1995-2014 기준선"),
    ("시점 (단기)", "2020-2039"),
    ("시점 (중기)", "2040-2059"),
    ("시점 (장기)", "2060-2079"),
    ("시점 (말기)", "2080-2099"),
]

r = 1
for label, val in info_rows:
    cell_style(ws, r, 1, label, bold=(label in ("항목","데이터 소스","기준")), bg=HEADER_BG if label in ("항목","데이터 소스","기준") else CARD_BG)
    cell_style(ws, r, 2, val,   bg=HEADER_BG if label in ("항목","데이터 소스","기준") else CARD_BG)
    ws.row_dimensions[r].height = 16
    r += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 55

# 임계점 테이블
THRESH_HEADERS = ["변수 코드", "변수명", "단위", "주의 기준 (Amber)", "위험 기준 (Red)", "임계점 근거 / 참조 기준"]
THRESH_DATA = [
    # CMIP6
    ("tasmax",            "최고기온",              "°C",           "> 33",      "> 38",      "ILO 열 노동안전: 33°C 야외작업 주의, 38°C 작업 중단 권고"),
    ("tasmin",            "최저기온",              "°C",           "> 22",      "—",         "22°C 이상 = 열대야 기준 (야간 냉각 불충분, 전력 야간 수요 증가)"),
    ("tas",               "평균기온",              "°C",           "> 25",      "> 30",      "25°C: 냉방 에너지 수요 급증; 30°C: 전력망 과부하 임계"),
    ("pr",                "강수량",                "mm/day",       "< 1 또는 > 10", "—",     "< 1 mm/day: 가뭄 징후; > 10 mm/day: 도시 침수 임계 강도"),
    ("sfcWind",           "지표풍속",              "m/s",          "> 10",      "> 15",      "10 m/s: 크레인·야외 작업 한계; 15 m/s: 구조물 손상 위험 (KBC 기준)"),
    # ETCCDI 열
    ("etccdi_wbgt",       "습구흑구온도 (WBGT)",   "°C",           "> 28",      "> 32",      "ILO/WBGT 기준: 28°C 경보(중강도 제한), 32°C 위험(경작업도 제한)"),
    ("etccdi_txx",        "월최고기온 극값 (TXx)", "°C",           "> 33",      "> 38",      "연중 피크 폭염 강도 — 설비 냉각·전력 임계 온도"),
    ("etccdi_su",         "여름일수 (SU>25°C)",    "days/yr",      "> 60",      "> 100",     "60일: 냉방 설계 기준; 100일: 아열대 기후 진입 수준"),
    ("etccdi_tr",         "열대야 (TR>20°C)",      "days/yr",      "> 10",      "> 30",      "10일: 야간 전력 수요 급증; 30일: 연속 야간 열 스트레스"),
    ("etccdi_wsdi",       "온난기간 (WSDI)",        "days/yr",      "> 10",      "> 30",      "연속 6일 이상 고온 지속 — 열파 사건 빈도 지표"),
    ("etccdi_fd",         "결빙일수 (FD<0°C)",     "days/yr",      "> 60",      "—",         "60일 이상: 동파·결빙 피해 위험 (한냉 기후 사업장)"),
    # ETCCDI 강수
    ("etccdi_cdd",        "연속건조일수 (CDD)",     "days",         "> 30",      "> 60",      "30일: 가뭄 주의; 60일: 심각한 취수 장애 (WMO 기준)"),
    ("etccdi_rx1day",     "1일 최대강수 (Rx1day)", "mm",           "> 50",      "> 100",     "50 mm: 도시 배수 한계; 100 mm: 홍수 피해 임계 강도"),
    ("etccdi_rx5day",     "5일 최대강수 (Rx5day)", "mm",           "> 100",     "> 200",     "댐·제방 설계 기준 초과 — 침수 사업 중단 위험"),
    ("etccdi_r95p",       "극한강수 (R95p)",        "mm/yr",        "> 200",     "> 500",     "연간 극한강수 누적량 — 장기 홍수·침식 리스크 지표"),
    ("etccdi_sdii",       "강수강도 (SDII)",        "mm/wet-day",   "> 20",      "—",         "20 mm/wet-day: 집중호우 패턴 (단기 침수 위험 증가)"),
    # PhyRisk
    ("heat_stress",       "열 스트레스",            "score 0-1",    "> 0.5",     "> 0.7",     "0.5: 노동생산성·설비 냉각 비용 유의미한 영향"),
    ("flood_risk",        "홍수 위험",              "score 0-1",    "> 0.5",     "> 0.7",     "시설 침수·사업 중단 가능성 (OS-Climate PhyRisk v2)"),
    ("cyclone_risk",      "사이클론 위험",          "score 0-1",    "> 0.5",     "> 0.7",     "풍속·해일 복합 피해 가능성 — 연안·항만 사업장 주의"),
    ("drought_risk",      "가뭄 위험",              "score 0-1",    "> 0.5",     "> 0.7",     "취수 안정성 위협 — 제조·에너지 냉각 공정 중단 위험"),
    ("wildfire_risk",     "산불 위험",              "score 0-1",    "> 0.5",     "> 0.7",     "산림 인접 사업장 화재 보험료·대피 계획 기준"),
    ("earthquake_risk",   "지진 위험",              "score 0-1",    "> 0.5",     "> 0.7",     "PhyRisk 지진 취약성 점수 — PSHA PGA와 교차 검증"),
    # CLIMADA EAL
    ("TC_EAL",            "태풍 연간기대손실 (EAL)", "score 0-100", "> 20",      "> 50",      "CLIMADA 역사 손실 지수: 20 = 보험 재검토; 50 = 중대 위험"),
    ("Flood_EAL",         "홍수 연간기대손실 (EAL)", "score 0-100", "> 20",      "> 50",      "연간기대손실(EAL) 지수: 홍수 역사 손실 기반 정량화"),
    ("Wildfire_EAL",      "산불 연간기대손실 (EAL)", "score 0-100", "> 20",      "> 50",      "산불 EAL: 산림 인접 사업장 손실 위험 정량화"),
    ("EQ_EAL",            "지진 연간기대손실 (EAL)", "score 0-100", "> 20",      "> 50",      "지진 EAL: 내진 설계 재점검 — CLIMADA OpenQuake 기반"),
    # Aqueduct
    ("aq_water_stress",   "수자원 스트레스",         "0-5",          "> 3",       "> 4",       "WRI Aqueduct: 3 = 높은 스트레스, 4 = 극도 스트레스 (CDP 공시 기준)"),
    ("aq_river_flood",    "하천홍수 위험",           "0-5",          "> 3",       "> 4",       "WRI: 하천 홍수 빈도·심도 복합 위험 지수"),
    ("aq_drought",        "가뭄 위험 (Aq)",          "0-5",          "> 3",       "> 4",       "WRI: 수자원 가뭄 취약성 지수 (농업·제조 취수 영향)"),
    # IBTrACS
    ("tc_annual_freq",    "태풍 연간 빈도",          "건/yr",        "> 0.5",     "> 1.0",     "0.5건: 주의 수준; 1.0건 이상: 연 1회 이상 태풍 직접 노출"),
    ("tc_max_wind_kt",    "최대 풍속",               "kt",           "> 64",      "> 96",      "64 kt = Cat 1 (구조 피해 시작); 96 kt = Cat 3 (중대 구조 피해)"),
    ("tc_cat3_count",     "Cat 3+ 태풍 건수",        "건",           "> 1",       "> 3",       "역사적 강태풍 직접 노출 건수 (1980~2023)"),
    # PSHA
    ("psha_pga_475",      "PGA 475년 재현주기",      "g",            "> 0.1",     "> 0.3",     "IBC/KBC 내진 설계: 0.1g 기본 설계값; 0.3g 고위험 (특수 설계 필요)"),
    ("psha_pga_2475",     "PGA 2475년 재현주기",     "g",            "> 0.2",     "> 0.5",     "최대 고려 지진(MCE): 0.2g 주의; 0.5g = 고위험 지역 특수 구조 기준"),
    # CCKP
    ("cckp_hi35",         "열지수 초과일 (HI>35°C)", "days/yr",      "> 20",      "> 60",      "20일: 냉방 인프라 유의미; 60일: 연중 냉방 필수 (열대 기후 임박)"),
    ("cckp_hd40",         "극한고온일 (Tmax>40°C)",  "days/yr",      "> 10",      "> 30",      "40°C 이상: 산업 냉각 시스템 설계 한계 초과 — 긴급 대응 필요"),
    ("cckp_tr26",         "열대야 한국기준 (>26°C)", "days/yr",      "> 10",      "> 30",      "한국 기상청 열대야 기준(26°C) — 야간 전력 수요 급증 지표"),
    ("cckp_cdd65",        "냉방도일 (CDD base 65°F)","°F·day/yr",   "> 1500",    "> 2500",    "에너지 인프라 냉방 설계 기준: 2500 이상 = 냉방 집약 지역"),
    ("cckp_hdd65",        "난방도일 (HDD base 65°F)","°F·day/yr",   "> 2000",    "> 4000",    "난방 에너지 수요: 4000 이상 = 한랭 기후 (가스·열 인프라 필수)"),
    ("cckp_spei12",       "가뭄지수 (SPEI-12)",      "index",        "< -0.5",    "< -1.5",    "SPEI 음수 = 건조: -0.5 경도 가뭄, -1.5 심각 가뭄 (FAO/WMO 기준)"),
    ("cckp_rxmonth",      "월최대강수 (Rx-month)",   "mm/month",     "> 250",     "> 400",     "250 mm: 하수·배수 설계 초과; 400 mm: 대규모 홍수 피해 기준"),
]

r += 2
# 임계점 섹션 제목
c = ws.cell(row=r, column=1, value="변수별 위험 임계점 (RAG 기준 근거)")
c.font      = hfont(bold=True, color=ACCENT, size=11)
c.fill      = hfill(HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.row_dimensions[r].height = 22
r += 1

# 임계점 컬럼 헤더
AMBER_COLOR = "FFB300"
RED_COLOR   = "C62828"
for ci, h in enumerate(THRESH_HEADERS, 1):
    c = ws.cell(row=r, column=ci, value=h)
    c.font      = hfont(bold=True, color="FFFFFF", size=9)
    c.fill      = hfill(HEADER_BG)
    c.border    = hborder()
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 16
r += 1

# 임계점 데이터 행
for i, (key, name_k, unit, amber, red, reason) in enumerate(THRESH_DATA):
    bg = CARD_BG if i % 2 == 0 else "0E1C30"
    vals = [key, name_k, unit, amber, red, reason]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.font      = hfont(bold=(ci == 1), color=ACCENT if ci == 1 else TEXT_FG, size=9)
        c.fill      = hfill(bg)
        c.border    = hborder()
        c.alignment = Alignment(horizontal="center" if ci <= 3 else "left", vertical="center", wrap_text=(ci == 6))
        if ci == 4 and amber not in ("—", ""):
            c.font = hfont(bold=True, color=AMBER_COLOR, size=9)
        if ci == 5 and red not in ("—", ""):
            c.font = hfont(bold=True, color=RED_COLOR, size=9)
    ws.row_dimensions[r].height = 28
    r += 1

# 임계점 열 너비
for ci, w in zip(range(1, 7), [18, 22, 13, 18, 18, 55]):
    ws.column_dimensions[get_column_letter(ci)].width = w

freeze(ws, "A3")


# ── 시트 7: 데이터 출처 안내 ──────────────────────────────────────────────────
ws = wb.create_sheet("데이터_출처_안내")
ws.sheet_view.showGridLines = False

SOURCE_META = [
    # (그룹명, 변수수, 제공기관, 출처·데이터셋, 해상도·커버리지,
    #  주요 용도, 주요 활용 분야, 비고)
    (
        "CMIP6\n기후변수",
        7,
        "PCMDI / WCRP\n(17개 글로벌 기후모델 컨소시엄)",
        "ESGF (Earth System Grid Federation)\n- MIROC6, MPI-ESM1-2-LR, IPSL-CM6A-LR 등 17종\n- SSP1-2.6 / SSP2-4.5 / SSP3-7.0 / SSP5-8.5",
        "전세계\n1°~2° 격자\n(약 100~200km)",
        "미래 기후 변화 기준값 제공\n온도·강수·바람·증발 시나리오의\n핵심 입력 데이터",
        "IPCC AR6 기반 기후 공시 (TCFD·ISSB S2)\nESG 투자자 기후 스트레스 테스트\n탄소 감축 경로 수립\n금융권 기후 리스크 내부 모델",
        "본 시스템의 베이스라인\n원시 CMIP6 값은 후속\n데이터로 고해상도 보완됨",
    ),
    (
        "ETCCDI\n극값 지수",
        13,
        "WMO / WCRP\nCCl & CLIVAR 공동 정의",
        "CMIP6 일별 데이터에서 산출\n- 온도 7종: TXx, TNn, SU, TR, FD, WSDI, WBGT\n- 강수 6종: CDD, CWD, Rx1day, Rx5day, R95p, SDII",
        "전세계\n(CMIP6 격자 동일)",
        "극한기후 사건 빈도·강도 정량화\n폭염·한파·극한강수 발생 패턴 분석\nWBGT로 열 노동 안전 평가",
        "보험·재보험 손실 모델링\n건물·인프라 설계 기준 (내열·내풍)\n공급망 사업 중단 리스크 평가\nILO 열 노동 안전 기준 적용",
        "ETCCDI 미존재 시\nCMIP6 변수로 회귀 추정\n(confidence 등급 제공)",
    ),
    (
        "PhyRisk\n(OS-Climate)",
        18,
        "OS-Climate\n(Linux Foundation 산하\n오픈소스 컨소시엄)",
        "OS-Climate PhyRisk v2\n- 18개 위험 유형 (0~1 스케일)\n- CMIP6 기반 물리 모델 구동",
        "전세계\n약 1/8° 격자\n(~14km)",
        "자산 수준 물리적 위험 점수 산출\n열·노동·홍수·사이클론·가뭄·\n지진·산불·산사태 18종 통합",
        "TCFD·ISSB S2 자산별 공시\nEU Taxonomy 기후 적응 적격성 평가\nK-녹색분류체계 (K-Taxonomy)\nPRI (책임투자원칙) 보고",
        "PhyRisk 데이터 부재 시\nCMIP6 변수로 추정값 자동 보완",
    ),
    (
        "Aqueduct\n수자원",
        6,
        "WRI\n(세계자원연구소)",
        "WRI Aqueduct 4.0\n- 수자원 스트레스, 홍수, 가뭄 6종\n- 현재 + 2030/2050 전망 포함\n- 0~5 스케일 (5=최고위험)",
        "전세계\n하천 유역 단위\n(~10km)",
        "제조·에너지 시설 수자원 위험 평가\n원수 취수 안정성 및 비용 리스크\n2050 수자원 스트레스 전망",
        "CDP 수자원 공시\n다국적기업 공급망 수자원 실사\n농업·음료·반도체 등 수자원 집약 산업\nIFC 환경·사회 기준(EHS Guidelines)",
        "정적 데이터\n(SSP·시점 무관,\n전 셀에 동일값)",
    ),
    (
        "IBTrACS\n태풍",
        3,
        "NOAA\n(미국 해양대기청)\nNational Centers for\nEnvironmental Information",
        "IBTrACS v04 (International Best Track\nArchive for Climate Stewardship)\n- 1980~2023년 역사적 태풍 전 수록\n- 빈도, 최대풍속, Cat3+ 태풍 집계",
        "전세계\n태풍 경로 격자\n통계",
        "역사적 태풍·사이클론 피해\n패턴 및 강도 기준선 제공\n사업장 주변 태풍 노출도 산정",
        "보험·재보험 PML(최대예상손실) 산정\n연안·항만 인프라 위험 분류\n공급망 사업장 위험 스크리닝\nECA (수출신용기관) 담보 평가",
        "역사 통계 기반 정적 데이터\n미래 태풍 강도 변화는\nPhyRisk cyclone_risk로 보완",
    ),
    (
        "PSHA\n지진재해",
        2,
        "GEM\n(Global Earthquake\nModel Foundation)\n/ USGS",
        "GEM OpenQuake Global Seismic Hazard\n- PGA 475년 재현주기 (10% 초과확률/50yr)\n- PGA 2475년 재현주기 (2% 초과확률/50yr)\n- 단위: g (중력가속도)",
        "전세계\n약 0.1° 격자",
        "지진 위험도 기준값 제공\n내진 설계 기준 적합성 판단\n지진 손실 추정(PGA → EAL 변환)",
        "IBC·KBC 내진 설계 등급 분류\n건물·플랜트 보험 인수 심사\n지진다발지역 사업 타당성 검토\nSendai Framework 재해위험 지표",
        "정적 데이터\nPGA → 리스크 점수 변환\n(pga_to_risk_score() 함수)",
    ),
    (
        "CLIMADA\nHDF5 EAL",
        4,
        "ETH Zürich\nWECC\n(CLIMADA 개발팀)\n/ 오픈소스",
        "CLIMADA (Climate Change Adaptation)\nHDF5 손실 모델 데이터베이스\n- TC_EAL: 태풍 연간기대손실 지수\n- Flood_EAL: 홍수 연간기대손실 지수\n- Wildfire_EAL: 산불 연간기대손실 지수\n- EQ_EAL: 지진 연간기대손실 지수\n- 역사적 관측 기반 손실 모델 (2000~2020년)",
        "전세계\n~30km 격자\n(역사적 정적값)",
        "정량 손실 모델 기반\n연간기대손실(EAL) 지수 산출\n자산 가치 대비 손실률 정량화\n보험·재보험 언더라이팅 핵심 입력\n타 점수형 데이터와 달리\n실제 손실액 기반 정량화",
        "재보험사 기후 손실 포트폴리오 관리\nMunich Re / Swiss Re 대재해 모델 비교\n국제개발은행 기후 사업 경제성 분석\n(BCR·NPV 기후 조정)\nIFRS 17 보험부채 기후 리스크 조정\nTCFD 시나리오 분석 정량화",
        "역사적 손실 기반 정적 데이터\nSSP·시점 무관 동일값\n미래 기후 변화에 따른 손실 변화는\nPhyRisk(cyclone/flood/wildfire/earthquake)\n로 보완하여 사용",
    ),
    (
        "CCKP\n에너지·극한열\n(World Bank)",
        13,
        "World Bank\n기후변화 지식 포털\n(CCKP)",
        "CCKP CMIP6 앙상블 0.25° 격자\n- 에너지 5종: HI>35, Tmax>40, TR>26, CDD, HDD\n- ETCCDI 교차검증 3종: CSDI, WSDI, CDD\n- 신규확장 5종: SPEI-12, GSL, HURS, ID, Rx-month",
        "전세계\n0.25° 격자\n(~28km)",
        "에너지 수요 변화(냉난방도일) 전망\n극한열·결빙일 부가 지수 산출\nCMIP6 ETCCDI와 상호 교차검증\n농업 생장기간·가뭄지수 제공",
        "에너지 인프라 계획 (전력망 용량 설계)\n건물 냉난방 에너지 수요 예측\n개발도상국 기후 적응 계획 (World Bank 사업)\n농업·식량안보 리스크 평가",
        "REST API로 실시간 쿼리\n추가 다운로드 없이\n임의 좌표 즉시 조회 가능",
    ),
    (
        "KMA_RDA\n(기상청\n행정구역)",
        6,
        "농촌진흥청 기상재해대응\n(weather.rda.go.kr)\n/ 기상청 NIMS",
        "ENS 앙상블 통계적 상세화\n- 한반도 167개 시·군·구 행정구역\n- SSP1-2.6 / SSP2-4.5 / SSP3-7.0 / SSP5-8.5\n- 변수: tasmax, tasmin, pr, sfcWind, hurs, rsds",
        "한반도\n167개 행정구역\n(약 20~30km)",
        "한국 국내 사업장 고해상도\n기후변화 시나리오 제공\n전국 행정구역별 기후 변화 비교",
        "국내 사업장 기후 물리적 리스크 공시\n한국형 녹색분류체계(K-Taxonomy)\nK-ESG 가이드라인 기후 공시\n지자체·공공기관 기후 적응계획 수립",
        "한반도 좌표 전용\n글로벌 사업장은 N/A\nCORDEX가 동일 변수 보완",
    ),
    (
        "CORDEX\n전역 RCM",
        7,
        "WCRP CORDEX\n(국제 공동 지역기후\n상세화 실험)\n주요: MOHC / GERICS",
        "ESGF CORDEX 전역 6개 도메인\n- EAS-22 동아시아 (~22km)\n- SEA-22 동남아 (~22km)\n- NAM-22 북미 (~22km)\n- EUR-11 유럽 (~11km)\n- SAM-44 남미 (~44km)\n- AFR-44 아프리카 (~44km)\n구동 GCM: MOHC-HadGEM2-ES / MPI-M-MPI-ESM-LR",
        "전세계 6개 도메인\n11~44km 해상도\n(CMIP6 대비 5~10배 고해상도)",
        "지역 규모 기후 변화 영향 정밀 평가\nCMIP6 전구 모델 대비\n지형·해양 효과 반영 고해상도 제공\n태풍·몬순·국지 강수 정밀화",
        "지역 인프라·시설 기후 취약성 평가\n도시 열섬·국지 홍수 위험 분석\nADB·World Bank 지역개발사업 기후 심사\n한국 K-ESG 글로벌 사업장 공시",
        "CMIP6와 동일 SSP 시나리오\n한반도는 KMA_RDA가 더\n높은 해상도로 우선 적용",
    ),
]

COL_WIDTHS = [16, 7, 22, 38, 18, 30, 35, 22]
COL_LABELS = ["그룹", "변수수", "제공 기관", "데이터 출처·데이터셋", "해상도·커버리지",
              "주요 용도", "주요 활용 분야", "비고"]

# 제목
r = 1
c = ws.cell(row=r, column=1, value="기후 물리적 리스크 — 데이터 출처 및 활용 안내")
c.font      = hfont(bold=True, color=ACCENT, size=12)
c.fill      = hfill(DARK_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COL_LABELS))
ws.row_dimensions[r].height = 24

# 컬럼 헤더
r = 2
header_row(ws, r, COL_LABELS)
ws.row_dimensions[r].height = 18

# 데이터 행
r = 3
for i, row_data in enumerate(SOURCE_META):
    bg = CARD_BG if i % 2 == 0 else "0E1C30"
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font      = hfont(bold=(ci == 1), color=ACCENT if ci == 1 else TEXT_FG, size=9)
        c.fill      = hfill(bg)
        c.border    = hborder()
        c.alignment = Alignment(horizontal="center" if ci == 2 else "left",
                                vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 90
    r += 1

# 열 너비
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 주석 행
r += 1
note = ("※ score(0~1): 0=무위험, 1=최고위험  |  0~5 스케일: WRI Aqueduct 기준  |  "
        "EAL score(0~100): CLIMADA 역사 손실 지수 (20 이상 주의·보험 재검토, 50 이상 위험)  |  "
        "정적 데이터: Aqueduct/IBTrACS/PSHA/CLIMADA는 SSP·시점 무관 동일값  |  "
        "KMA·CORDEX: 한반도 좌표 = 고해상도값, 해외 좌표 = KMA N/A / CORDEX 해당 도메인값")
c = ws.cell(row=r, column=1, value=note)
c.font      = hfont(bold=False, color=DIM_FG, size=8)
c.fill      = hfill(DARK_BG)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COL_LABELS))
ws.row_dimensions[r].height = 30

freeze(ws, "A3")


# ── 저장 ──────────────────────────────────────────────────────────────────────
safe_name = name.replace(" ", "_").replace("/", "-")
out_path = OUT_DIR / f"{safe_name}_{lat}_{lon}_climate_risk.xlsx"
wb.save(str(out_path))
logger.info(f"\n✓ 저장 완료: {out_path}")
print(f"\n파일 경로: {out_path}")
